import logging
from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import IntegrityError
from fpdf import FPDF
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os, re, tempfile, json
from io import BytesIO

app = Flask(__name__)

# ---------------------------------------------------------------------------
# Database configuration: PostgreSQL via DATABASE_URL env var, or SQLite fallback
# ---------------------------------------------------------------------------
database_url = os.environ.get("DATABASE_URL")
if database_url and database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
if not database_url:
    database_url = "sqlite:///invoice.db"
app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


class Invoice(db.Model):
    __tablename__ = "invoices"
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    invoice_num = db.Column(db.String, unique=True, nullable=False, default="")
    date = db.Column(db.String, nullable=False, default="")
    client_name = db.Column(db.String, nullable=False, default="")
    client_address = db.Column(db.String, nullable=False, default="")
    total = db.Column(db.Float, nullable=False, default=0.0)
    remise = db.Column(db.Float, nullable=False, default=0.0)
    payer = db.Column(db.Float, nullable=False, default=0.0)
    net_total = db.Column(db.Float, nullable=False, default=0.0)
    items = db.Column(db.JSON, nullable=False, default=list)
    created_at = db.Column(
        db.String,
        default=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )


# Create all tables on startup (runs at import time for Gunicorn,
# and also when executed directly via python app.py).
with app.app_context():
    db.create_all()


COMPANY_NAME = "Votre Société"
COMPANY_ADDRESS = "Rue Tange Center Al hirafiyin N 25 Imzouren Al Hoceima"
COMPANY_PHONE = ""
COMPANY_EMAIL = ""


GOLD = (201, 168, 76)
DARK = (51, 51, 51)
GRAY_BG = (240, 240, 240)
GRAY_TEXT = (136, 136, 136)
MID_GRAY = (85, 85, 85)


def save_invoice(data):
    """
    Create or update an invoice.

    Edit mode:   data contains 'id'  → UPDATE the existing row keeping its id.
    Create mode: data has no 'id'    → INSERT a new row; the DB assigns a unique
                                       auto-incremented id.
    """
    invoice_num    = data.get("invoice_num", "")
    date_val       = data.get("date", "")
    client_name    = data.get("client_name", "")
    client_address = data.get("client_address", "")
    total          = data.get("total", 0.0)
    remise         = data.get("remise", 0.0)
    payer          = data.get("payer", 0.0)
    net_total      = data.get("net_total", 0.0)
    items          = data.get("items", [])

    existing_id = data.get("id")

    # Enforce unique invoice numbers
    dup_query = Invoice.query.filter(Invoice.invoice_num == invoice_num)
    if existing_id:
        dup_query = dup_query.filter(Invoice.id != existing_id)
    if dup_query.first():
        raise ValueError("Ce numéro de facture existe déjà.")

    if existing_id:
        invoice = db.session.get(Invoice, existing_id)
        if invoice:
            invoice.invoice_num = invoice_num
            invoice.date = date_val
            invoice.client_name = client_name
            invoice.client_address = client_address
            invoice.total = total
            invoice.remise = remise
            invoice.payer = payer
            invoice.net_total = net_total
            invoice.items = items
        invoice_id = existing_id
    else:
        invoice = Invoice(
            invoice_num=invoice_num,
            date=date_val,
            client_name=client_name,
            client_address=client_address,
            total=total,
            remise=remise,
            payer=payer,
            net_total=net_total,
            items=items
        )
        db.session.add(invoice)
        db.session.flush()
        invoice_id = invoice.id

    db.session.commit()
    return invoice_id


class InvoicePDF(FPDF):
    def header(self):
        if self.page_no() > 1:
            return

        self.set_y(25)

        logo_path = os.path.join(os.path.dirname(__file__), "static", "logo.png")
        if os.path.exists(logo_path):
            self.image(logo_path, x=self.l_margin, y=14, w=48)

        self.set_font("Helvetica", "B", 20)
        self.set_text_color(*GOLD)
        self.cell(0, 8, "COCINA ESPAÑOLA", align="R", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "I", 14)
        self.cell(0, 6, "Art MDF", align="R", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 9)
        self.set_text_color(*MID_GRAY)
        self.cell(0, 5, "Rue Tange Center Al hirafiyin N 25", align="R", new_x="LMARGIN", new_y="NEXT")
        self.cell(0, 5, "Imzouren AL Hoceima", align="R", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "B", 9)
        self.set_text_color(*DARK)
        client_name = getattr(self, "_client_name", "")
        self.cell(0, 5, f"Client: {client_name}", align="R", new_x="LMARGIN", new_y="NEXT")
        self.ln(10)

        self.set_draw_color(*GOLD)
        self.set_line_width(0.6)
        self.set_font("Helvetica", "B", 18)
        self.set_text_color(*DARK)
        facture_num = getattr(self, "_facture_num", "")
        show_fn = getattr(self, "_show_facture_num", False)
        title = "FACTURE"
        if show_fn and facture_num:
            title += f" N\u00b0 {facture_num}"
        self.cell(0, 10, title, align="C", new_x="LMARGIN", new_y="NEXT")
        y = self.get_y()
        self.line(95, y, 115, y)
        self.ln(8)

        date_val = getattr(self, "_date", "")
        self.set_font("Helvetica", "", 10)
        self.set_text_color(*MID_GRAY)
        self.cell(0, 6, f"Date de facture: {date_val}", align="R", new_x="LMARGIN", new_y="NEXT")
        self.ln(6)

    def footer(self):
        self.set_y(-25)
        self.set_draw_color(200, 200, 200)
        self.set_line_width(0.3)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.ln(4)

        self.set_font("Helvetica", "", 7)
        self.set_text_color(*GRAY_TEXT)
        page_w = self.w - self.l_margin - self.r_margin
        col_w = page_w / 3
        self.set_x(self.l_margin)
        self.cell(col_w, 4, "Rue Tange Center Al hirafiyin N 25", align="C")
        self.cell(col_w, 4, "Tel: +212 6 71 68 75 98", align="C")
        self.cell(col_w, 4, "Instagram: cocinaespanola", align="C")
        self.ln()
        self.set_x(self.l_margin)
        self.cell(col_w, 4, "Imzouren AL Hoceima", align="C")
        self.cell(col_w, 4, "", align="C")
        self.set_text_color(*DARK)
        self.cell(col_w, 4, f"{self.page_no()}/{{nb}}", align="R")


def safe_filename(name):
    name = re.sub(r'[<>:"/\\|?*]', '_', name).strip()
    if not name:
        name = "facture"
    return name


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/generate_pdf", methods=["POST"])
def generate_pdf():
    data = request.get_json()

    pdf = InvoicePDF()
    pdf._client_name = data.get("client_name", "")
    pdf._date = data.get("date", "")
    pdf._facture_num = data.get("invoice_num", "")
    pdf._show_facture_num = data.get("show_facture_num", False)
    pdf.set_auto_page_break(auto=True, margin=30)
    pdf.add_page()

    col_w = [72, 28, 44, 38]
    headers = ["Description", "Quantité", "Prix DH", "Montant DH"]

    def draw_table_header():
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(*GRAY_BG)
        pdf.set_text_color(*DARK)
        pdf.set_draw_color(200, 200, 200)
        pdf.set_draw_color(0, 0, 0)
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 8, h, border=1, align="C", fill=True)
        pdf.ln()

    draw_table_header()

    pdf.set_font("Helvetica", "", 9)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(*DARK)
    last_page = pdf.page_no()
    for item in data['items']:
        if pdf.page_no() != last_page:
            draw_table_header()
            last_page = pdf.page_no()
        pdf.cell(col_w[0], 7, item['description'], border=1)
        pdf.cell(col_w[1], 7, str(item['quantity']), border=1, align="C")
        pdf.cell(col_w[2], 7, f"{item['unit_price']:,.2f}".replace(',', ' '), border=1, align="C")
        pdf.cell(col_w[3], 7, f"{item['total']:,.2f}".replace(',', ' '), border=1, align="C")
        pdf.ln()

    # Total
    total = data['total']
    remise = data.get('remise', 0)
    payer = data.get('payer', 0)
    net_total = data.get('net_total', total)

    pdf.ln(4)
    total_x = pdf.l_margin + col_w[0] + col_w[1]

    pdf.set_draw_color(0, 0, 0)

    def price_line(label, value, val_color=None, bold=False):
        pdf.set_x(total_x)
        pdf.set_fill_color(*GRAY_BG)
        pdf.set_draw_color(0, 0, 0)
        pdf.set_font("Helvetica", "B" if bold else "", 10)
        if val_color:
            pdf.set_text_color(*val_color)
        else:
            pdf.set_text_color(*DARK)
        pdf.cell(col_w[2], 10, label, border=1, align="C", fill=True)
        pdf.cell(col_w[3], 10, value, border=1, align="C", fill=True)
        pdf.ln(10)

    if remise > 0 or payer > 0:
        price_line("Total en DH", f"{total:,.2f} DH".replace(',', ' '))
    if remise > 0:
        price_line("Remise en DH", f"-{remise:,.2f} DH".replace(',', ' '))
    if payer > 0:
        price_line("Payer en DH", f"-{payer:,.2f} DH".replace(',', ' '))
    price_line("Total a Payer en DH", f"{net_total:,.2f} DH".replace(',', ' '), val_color=GOLD, bold=True)

    name = f"{safe_filename(data['client_name'])}.pdf"
    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=name, mimetype="application/pdf")


@app.route("/generate_excel", methods=["POST"])
def generate_excel():
    data = request.get_json()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facture"

    GOLD_HEX = "C9A84C"
    GRAY_HEX = "F0F0F0"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color=GRAY_HEX, end_color=GRAY_HEX, fill_type="solid")

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # Header
    ws.merge_cells('A1:C1')
    ws['A1'] = "COCINA ESPAÑOLA"
    ws['A1'].font = Font(name="Arial", bold=True, size=20, color=GOLD_HEX)

    ws.merge_cells('A2:C2')
    ws['A2'] = "Art MDF"
    ws['A2'].font = Font(name="Arial", italic=True, size=14, color=GOLD_HEX)

    ws.merge_cells('D1:D2')
    ws['D1'] = "Rue Tange Center Al hirafiyin N 25"
    ws['D1'].font = Font(name="Arial", size=9, color="555555")
    ws['D1'].alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('D3:D3')
    ws['D3'] = "Imzouren AL Hoceima"
    ws['D3'].font = Font(name="Arial", size=9, color="555555")
    ws['D3'].alignment = Alignment(horizontal='right')

    ws['C4'] = "Client:"
    ws['C4'].font = Font(name="Arial", bold=True, size=10, color="333333")
    ws['C4'].alignment = Alignment(horizontal='right')

    ws['D4'] = data.get("client_name", "")
    ws['D4'].font = Font(name="Arial", bold=True, size=10, color="333333")
    ws['D4'].alignment = Alignment(horizontal='right')

    # Title
    ws.merge_cells('A6:D6')
    title = "FACTURE"
    show_fn = data.get("show_facture_num", False)
    fn = data.get("invoice_num", "")
    if show_fn and fn:
        title += f" N\u00b0 {fn}"
    ws['A6'] = title
    ws['A6'].font = Font(name="Arial", bold=True, size=18, color="333333")
    ws['A6'].alignment = Alignment(horizontal='center')

    # Gold underline (row 7) - use bottom border on row 6 instead
    ws['A6'].border = Border(bottom=Side(style='medium', color=GOLD_HEX))

    # Date
    ws.merge_cells('C8:D8')
    ws['C8'] = f"Date de facture: {data.get('date', '')}"
    ws['C8'].font = Font(name="Arial", size=10, color="555555")
    ws['C8'].alignment = Alignment(horizontal='right')

    # Table header
    headers = ['Description', 'Quantité', 'Prix DH', 'Montant DH']
    header_row = 10
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="333333")
        cell.fill = gray_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Table rows
    for i, item in enumerate(data['items']):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=item['description']).border = thin_border
        ws.cell(row=r, column=1).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=2, value=item['quantity']).border = thin_border
        ws.cell(row=r, column=2).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=3, value=item['unit_price']).border = thin_border
        ws.cell(row=r, column=3).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=4, value=item['total']).border = thin_border
        ws.cell(row=r, column=4).font = Font(name="Arial", size=10)
        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')

    # Total lines
    total = data['total']
    remise = data.get('remise', 0)
    payer = data.get('payer', 0)
    net_total = data.get('net_total', total)

    r = header_row + 1 + len(data['items']) + 1

    if remise > 0 or payer > 0:
        # Total en DH
        ws.cell(row=r, column=3, value="Total en DH")
        ws.cell(row=r, column=3).font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=total)
        ws.cell(row=r, column=4).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    if remise > 0:
        # Remise en DH
        ws.cell(row=r, column=3, value="Remise en DH")
        ws.cell(row=r, column=3).font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=-remise)
        ws.cell(row=r, column=4).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    if payer > 0:
        # Payer en DH
        ws.cell(row=r, column=3, value="Payer en DH")
        ws.cell(row=r, column=3).font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=-payer)
        ws.cell(row=r, column=4).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    # Total a Payer en DH
    ws.cell(row=r, column=3, value="Total à Payer en DH")
    ws.cell(row=r, column=3).font = Font(name="Arial", bold=True, size=11, color="333333")
    ws.cell(row=r, column=3).fill = gray_fill
    ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=3).border = thin_border
    ws.cell(row=r, column=4, value=net_total)
    ws.cell(row=r, column=4).font = Font(name="Arial", bold=True, size=12, color=GOLD_HEX)
    ws.cell(row=r, column=4).fill = gray_fill
    ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=4).border = thin_border

    # Footer
    footer_row = r + 2
    ws.merge_cells(f'A{footer_row}:A{footer_row}')
    ws[f'A{footer_row}'] = "Rue Tange Center Al hirafiyin N 25"
    ws[f'A{footer_row}'].font = Font(name="Arial", size=8, color="888888")
    ws[f'A{footer_row}'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'B{footer_row}:C{footer_row}')
    ws[f'B{footer_row}'] = "Tel: +212 6 71 68 75 98"
    ws[f'B{footer_row}'].font = Font(name="Arial", size=8, color="888888")
    ws[f'B{footer_row}'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'D{footer_row}:D{footer_row}')
    ws[f'D{footer_row}'] = "Instagram: cocinaespanola"
    ws[f'D{footer_row}'].font = Font(name="Arial", size=8, color="888888")
    ws[f'D{footer_row}'].alignment = Alignment(horizontal='center')

    addr_row = footer_row + 1
    ws.merge_cells(f'A{addr_row}:A{addr_row}')
    ws[f'A{addr_row}'] = "Imzouren AL Hoceima"
    ws[f'A{addr_row}'].font = Font(name="Arial", size=8, color="888888")
    ws[f'A{addr_row}'].alignment = Alignment(horizontal='center')

    base = safe_filename(data['client_name'])
    name = f"{base}.xlsx"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(tmp.name, as_attachment=True, download_name=name)


@app.route("/preview", methods=["POST"])
def preview():
    data = request.get_json()
    return render_template("invoice_pdf.html", data=data, company_name=COMPANY_NAME, company_address=COMPANY_ADDRESS)


@app.route("/save_invoice", methods=["POST"])
def save_invoice_route():
    data = request.get_json()
    try:
        invoice_id = save_invoice(data)
        if invoice_id is None:
            return jsonify({"error": "Facture non trouvée"}), 404
        return jsonify({"id": invoice_id})
    except ValueError as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 400
    except IntegrityError:
        db.session.rollback()
        app.logger.exception("IntegrityError saving invoice")
        return jsonify({"error": "Numéro de facture déjà utilisé."}), 400
    except Exception as e:
        db.session.rollback()
        app.logger.exception("Error saving invoice")
        return jsonify({"error": f"Erreur lors de l'enregistrement de la facture: {str(e)}"}), 500


def get_invoice_by_id(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        return None
    return {
        "id": invoice.id,
        "invoice_num": invoice.invoice_num,
        "date": invoice.date,
        "client_name": invoice.client_name,
        "client_address": invoice.client_address,
        "total": invoice.total,
        "remise": invoice.remise,
        "payer": invoice.payer,
        "net_total": invoice.net_total,
        "items": invoice.items
    }


@app.route("/invoices/<int:invoice_id>/json")
def get_invoice_json(invoice_id):
    """Return a single invoice as JSON — used by the frontend edit form."""
    data = get_invoice_by_id(invoice_id)
    if not data:
        return jsonify({"error": "Facture non trouvée"}), 404
    return jsonify(data)


@app.route("/invoices")
def invoices_list():
    try:
        rows = Invoice.query.order_by(Invoice.date.desc(), Invoice.id.desc()).all()
        invoices = [
            {
                "id": inv.id,
                "invoice_num": inv.invoice_num,
                "date": inv.date,
                "client_name": inv.client_name,
                "net_total": inv.net_total
            }
            for inv in rows
        ]
    except Exception:
        invoices = []
    return render_template("invoices.html", invoices=invoices)


@app.route("/invoices/<int:invoice_id>/pdf")
def download_saved_pdf(invoice_id):
    data = get_invoice_by_id(invoice_id)
    if not data:
        return "Facture non trouvée", 404

    pdf = InvoicePDF()
    pdf._client_name = data.get("client_name", "")
    pdf._date = data.get("date", "")
    pdf._facture_num = data.get("invoice_num", "")
    pdf._show_facture_num = True
    pdf.set_auto_page_break(auto=True, margin=30)
    pdf.add_page()

    col_w = [72, 28, 44, 38]
    headers = ["Description", "Quantité", "Prix DH", "Montant DH"]

    def draw_table_header():
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(*GRAY_BG)
        pdf.set_text_color(*DARK)
        pdf.set_draw_color(0, 0, 0)
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 8, h, border=1, align="C", fill=True)
        pdf.ln()

    draw_table_header()

    pdf.set_font("Helvetica", "", 9)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(*DARK)
    last_page = pdf.page_no()
    for item in data['items']:
        if pdf.page_no() != last_page:
            draw_table_header()
            last_page = pdf.page_no()
        pdf.cell(col_w[0], 7, item['description'], border=1)
        pdf.cell(col_w[1], 7, str(item['quantity']), border=1, align="C")
        pdf.cell(col_w[2], 7, f"{item['unit_price']:,.2f}".replace(',', ' '), border=1, align="C")
        pdf.cell(col_w[3], 7, f"{item['total']:,.2f}".replace(',', ' '), border=1, align="C")
        pdf.ln()

    # Total
    total = data['total']
    remise = data.get('remise', 0)
    payer = data.get('payer', 0)
    net_total = data.get('net_total', total)

    pdf.ln(4)
    total_x = pdf.l_margin + col_w[0] + col_w[1]

    pdf.set_draw_color(0, 0, 0)

    def price_line(label, value, val_color=None, bold=False):
        pdf.set_x(total_x)
        pdf.set_fill_color(*GRAY_BG)
        pdf.set_draw_color(0, 0, 0)
        pdf.set_font("Helvetica", "B" if bold else "", 10)
        if val_color:
            pdf.set_text_color(*val_color)
        else:
            pdf.set_text_color(*DARK)
        pdf.cell(col_w[2], 10, label, border=1, align="C", fill=True)
        pdf.cell(col_w[3], 10, value, border=1, align="C", fill=True)
        pdf.ln(10)

    if remise > 0 or payer > 0:
        price_line("Total en DH", f"{total:,.2f} DH".replace(',', ' '))
    if remise > 0:
        price_line("Remise en DH", f"-{remise:,.2f} DH".replace(',', ' '))
    if payer > 0:
        price_line("Payer en DH", f"-{payer:,.2f} DH".replace(',', ' '))
    price_line("Total a Payer en DH", f"{net_total:,.2f} DH".replace(',', ' '), val_color=GOLD, bold=True)

    name = f"{safe_filename(data['client_name'])}.pdf"
    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=name, mimetype="application/pdf")


@app.route("/invoices/<int:invoice_id>/excel")
def download_saved_excel(invoice_id):
    data = get_invoice_by_id(invoice_id)
    if not data:
        return "Facture non trouvée", 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facture"

    GOLD_HEX = "C9A84C"
    GRAY_HEX = "F0F0F0"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color=GRAY_HEX, end_color=GRAY_HEX, fill_type="solid")

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # Header
    ws.merge_cells('A1:C1')
    ws['A1'] = "COCINA ESPAÑOLA"
    ws['A1'].font = Font(name="Arial", bold=True, size=20, color=GOLD_HEX)

    ws.merge_cells('A2:C2')
    ws['A2'] = "Art MDF"
    ws['A2'].font = Font(name="Arial", italic=True, size=14, color=GOLD_HEX)

    ws.merge_cells('D1:D2')
    ws['D1'] = "Rue Tange Center Al hirafiyin N 25"
    ws['D1'].font = Font(name="Arial", size=9, color="555555")
    ws['D1'].alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('D3:D3')
    ws['D3'] = "Imzouren AL Hoceima"
    ws['D3'].font = Font(name="Arial", size=9, color="555555")
    ws['D3'].alignment = Alignment(horizontal='right')

    ws['C4'] = "Client:"
    ws['C4'].font = Font(name="Arial", bold=True, size=10, color="333333")
    ws['C4'].alignment = Alignment(horizontal='right')

    ws['D4'] = data.get("client_name", "")
    ws['D4'].font = Font(name="Arial", bold=True, size=10, color="333333")
    ws['D4'].alignment = Alignment(horizontal='right')

    # Title
    ws.merge_cells('A6:D6')
    title = "FACTURE"
    fn = data.get("invoice_num", "")
    if fn:
        title += f" N\u00b0 {fn}"
    ws['A6'] = title
    ws['A6'].font = Font(name="Arial", bold=True, size=18, color="333333")
    ws['A6'].alignment = Alignment(horizontal='center')

    # Gold underline (row 7) - use bottom border on row 6 instead
    ws['A6'].border = Border(bottom=Side(style='medium', color=GOLD_HEX))

    # Date
    ws.merge_cells('C8:D8')
    ws['C8'] = f"Date de facture: {data.get('date', '')}"
    ws['C8'].font = Font(name="Arial", size=10, color="555555")
    ws['C8'].alignment = Alignment(horizontal='right')

    # Table header
    headers = ['Description', 'Quantité', 'Prix DH', 'Montant DH']
    header_row = 10
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="333333")
        cell.fill = gray_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Table rows
    for i, item in enumerate(data['items']):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=item['description']).border = thin_border
        ws.cell(row=r, column=1).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=2, value=item['quantity']).border = thin_border
        ws.cell(row=r, column=2).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=3, value=item['unit_price']).border = thin_border
        ws.cell(row=r, column=3).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=4, value=item['total']).border = thin_border
        ws.cell(row=r, column=4).font = Font(name="Arial", size=10)
        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')

    # Total lines
    total = data['total']
    remise = data.get('remise', 0)
    payer = data.get('payer', 0)
    net_total = data.get('net_total', total)

    r = header_row + 1 + len(data['items']) + 1

    if remise > 0 or payer > 0:
        # Total en DH
        ws.cell(row=r, column=3, value="Total en DH")
        ws.cell(row=r, column=3).font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=total)
        ws.cell(row=r, column=4).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    if remise > 0:
        # Remise en DH
        ws.cell(row=r, column=3, value="Remise en DH")
        ws.cell(row=r, column=3).font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=-remise)
        ws.cell(row=r, column=4).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    if payer > 0:
        # Payer en DH
        ws.cell(row=r, column=3, value="Payer en DH")
        ws.cell(row=r, column=3).font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=-payer)
        ws.cell(row=r, column=4).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    # Total a Payer en DH
    ws.cell(row=r, column=3, value="Total à Payer en DH")
    ws.cell(row=r, column=3).font = Font(name="Arial", bold=True, size=11, color="333333")
    ws.cell(row=r, column=3).fill = gray_fill
    ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=3).border = thin_border
    ws.cell(row=r, column=4, value=net_total)
    ws.cell(row=r, column=4).font = Font(name="Arial", bold=True, size=12, color=GOLD_HEX)
    ws.cell(row=r, column=4).fill = gray_fill
    ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=4).border = thin_border

    # Footer
    footer_row = r + 2
    ws.merge_cells(f'A{footer_row}:A{footer_row}')
    ws[f'A{footer_row}'] = "Rue Tange Center Al hirafiyin N 25"
    ws[f'A{footer_row}'].font = Font(name="Arial", size=8, color="888888")
    ws[f'A{footer_row}'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'B{footer_row}:C{footer_row}')
    ws[f'B{footer_row}'] = "Tel: +212 6 71 68 75 98"
    ws[f'B{footer_row}'].font = Font(name="Arial", size=8, color="888888")
    ws[f'B{footer_row}'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'D{footer_row}:D{footer_row}')
    ws[f'D{footer_row}'] = "Instagram: cocinaespanola"
    ws[f'D{footer_row}'].font = Font(name="Arial", size=8, color="888888")
    ws[f'D{footer_row}'].alignment = Alignment(horizontal='center')

    addr_row = footer_row + 1
    ws.merge_cells(f'A{addr_row}:A{addr_row}')
    ws[f'A{addr_row}'] = "Imzouren AL Hoceima"
    ws[f'A{addr_row}'].font = Font(name="Arial", size=8, color="888888")
    ws[f'A{addr_row}'].alignment = Alignment(horizontal='center')

    base = safe_filename(data['client_name'])
    name = f"{base}.xlsx"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(tmp.name, as_attachment=True, download_name=name)


@app.route("/invoices/<int:invoice_id>/delete", methods=["POST"])
def delete_invoice_route(invoice_id):
    try:
        invoice = db.session.get(Invoice, invoice_id)
        if invoice:
            db.session.delete(invoice)
            db.session.commit()
        return jsonify({"success": True})
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Erreur lors de la suppression."}), 500


if __name__ == "__main__":
    os.makedirs("generated", exist_ok=True)
    app.run(debug=True, use_reloader=False)
