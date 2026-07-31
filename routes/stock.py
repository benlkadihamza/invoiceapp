import os
import uuid
import tempfile
from datetime import datetime
from io import BytesIO
from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from flask_login import login_required
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from models import db, Product, StockHistory

stock_bp = Blueprint('stock', __name__, url_prefix='/stock')

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}

GOLD = (201, 168, 76)
DARK = (51, 51, 51)
GRAY_BG = (240, 240, 240)
GRAY_TEXT = (136, 136, 136)
MID_GRAY = (85, 85, 85)


def allowed_image(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def get_upload_folder():
    folder = os.path.join(current_app.static_folder, "uploads", "products")
    os.makedirs(folder, exist_ok=True)
    return folder


def stock_status_from_db(products):
    rows = []
    for p in products:
        qty = p.stock_quantity
        if qty == 0:
            label = "FINISHED"
        elif qty <= 5:
            label = "LOW STOCK"
        else:
            label = "IN STOCK"
        rows.append({
            "title": p.title,
            "stock_quantity": qty,
            "status": label,
        })
    return rows


def _get_stock_status_label(qty):
    if qty == 0:
        return "OUT OF STOCK"
    if qty < 5:
        return "LOW STOCK"
    return "IN STOCK"


def _get_stock_status_class(qty):
    if qty == 0:
        return "out-of-stock"
    if qty < 5:
        return "low-stock"
    return "in-stock"


def _build_stock_status_data():
    products = Product.query.order_by(Product.title.asc()).all()
    total_products = len(products)
    total_stock = sum(p.stock_quantity for p in products)
    low_stock = sum(1 for p in products if 0 < p.stock_quantity < 5)
    out_of_stock = sum(1 for p in products if p.stock_quantity == 0)
    rows = []
    for p in products:
        rows.append({
            "id": p.id,
            "title": p.title,
            "photo": p.photo,
            "stock_quantity": p.stock_quantity,
            "updated_at": p.updated_at or "",
            "status": _get_stock_status_label(p.stock_quantity),
            "status_class": _get_stock_status_class(p.stock_quantity),
        })
    return {
        "rows": rows,
        "total_products": total_products,
        "total_stock": total_stock,
        "low_stock": low_stock,
        "out_of_stock": out_of_stock,
    }


@stock_bp.route('')
@stock_bp.route('/')
@login_required
def stock_page():
    products = Product.query.order_by(Product.title.asc()).all()
    total_products = len(products)
    total_stock = sum(p.stock_quantity for p in products)
    low_stock = sum(1 for p in products if 0 < p.stock_quantity < 5)
    out_of_stock = sum(1 for p in products if p.stock_quantity == 0)
    return render_template(
        "stock.html", products=products,
        total_products=total_products, total_stock=total_stock,
        low_stock=low_stock, out_of_stock=out_of_stock,
    )


from idempotency import idempotent_route


@stock_bp.route('/add', methods=['POST'])
@login_required
@idempotent_route()
def stock_add_product():
    title = request.form.get("title", "").strip()
    if not title:
        return jsonify({"error": "Le nom du produit est requis."}), 400
    if Product.query.filter_by(title=title).first():
        return jsonify({"error": "Un produit avec ce nom existe déjà."}), 400
    try:
        photo = ""
        if "photo" in request.files:
            f = request.files["photo"]
            if f.filename and allowed_image(f.filename):
                ext = f.filename.rsplit(".", 1)[1].lower()
                filename = f"{uuid.uuid4().hex}.{ext}"
                upload_folder = get_upload_folder()
                f.save(os.path.join(upload_folder, filename))
                photo = filename
        product = Product(title=title, photo=photo, stock_quantity=0)
        db.session.add(product)
        db.session.commit()
        return jsonify({"success": True, "id": product.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erreur lors de l'ajout: {str(e)}"}), 500


@stock_bp.route('/edit/<int:product_id>', methods=['POST'])
@login_required
@idempotent_route()
def stock_edit_product(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        return jsonify({"error": "Produit non trouvé."}), 404
    title = request.form.get("title", "").strip()
    if not title:
        return jsonify({"error": "Le nom du produit est requis."}), 400
    existing = Product.query.filter(
        Product.title == title, Product.id != product_id
    ).first()
    if existing:
        return jsonify({"error": "Un produit avec ce nom existe déjà."}), 400
    try:
        upload_folder = get_upload_folder()
        if "photo" in request.files:
            f = request.files["photo"]
            if f.filename and allowed_image(f.filename):
                ext = f.filename.rsplit(".", 1)[1].lower()
                filename = f"{uuid.uuid4().hex}.{ext}"
                f.save(os.path.join(upload_folder, filename))
                if product.photo:
                    old_path = os.path.join(upload_folder, product.photo)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                product.photo = filename
        product.title = title
        product.updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erreur lors de la modification: {str(e)}"}), 500


@stock_bp.route('/delete/<int:product_id>', methods=['POST'])
@login_required
@idempotent_route()
def stock_delete_product(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        return jsonify({"error": "Produit non trouvé."}), 404
    try:
        upload_folder = get_upload_folder()
        if product.photo:
            photo_path = os.path.join(upload_folder, product.photo)
            if os.path.exists(photo_path):
                os.remove(photo_path)
        StockHistory.query.filter_by(product_id=product_id).delete()
        db.session.delete(product)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erreur lors de la suppression: {str(e)}"}), 500


@stock_bp.route('/add-stock/<int:product_id>', methods=['POST'])
@login_required
@idempotent_route()
def stock_add_stock(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        return jsonify({"error": "Produit non trouvé."}), 404
    data = request.get_json(silent=True) or {}
    qty = data.get("quantity")
    if qty is None:
        return jsonify({"error": "Quantité requise."}), 400
    try:
        qty = int(qty)
    except (TypeError, ValueError):
        return jsonify({"error": "Quantité invalide."}), 400
    if qty < 1:
        return jsonify({"error": "La quantité doit être au moins 1."}), 400
    try:
        stock_before = product.stock_quantity
        product.stock_quantity += qty
        product.updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        history = StockHistory(
            product_id=product.id, change_type="ADD", quantity=qty,
            stock_before=stock_before, stock_after=product.stock_quantity,
        )
        db.session.add(history)
        db.session.commit()
        return jsonify({"success": True, "stock": product.stock_quantity})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erreur lors de l'ajout de stock: {str(e)}"}), 500


@stock_bp.route('/remove-stock/<int:product_id>', methods=['POST'])
@login_required
@idempotent_route()
def stock_remove_stock(product_id):
    product = db.session.get(Product, product_id)
    if not product:
        return jsonify({"error": "Produit non trouvé."}), 404
    data = request.get_json(silent=True) or {}
    qty = data.get("quantity")
    if qty is None:
        return jsonify({"error": "Quantité requise."}), 400
    try:
        qty = int(qty)
    except (TypeError, ValueError):
        return jsonify({"error": "Quantité invalide."}), 400
    if qty < 1:
        return jsonify({"error": "La quantité doit être au moins 1."}), 400
    if qty > product.stock_quantity:
        return jsonify({
            "error": f"Stock insuffisant. Stock disponible: {product.stock_quantity}"
        }), 400
    try:
        stock_before = product.stock_quantity
        product.stock_quantity -= qty
        product.updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        history = StockHistory(
            product_id=product.id, change_type="REMOVE", quantity=qty,
            stock_before=stock_before, stock_after=product.stock_quantity,
        )
        db.session.add(history)
        db.session.commit()
        return jsonify({"success": True, "stock": product.stock_quantity})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erreur lors du retrait de stock: {str(e)}"}), 500


@stock_bp.route('/report')
@login_required
def stock_report():
    now = datetime.now()
    month = request.args.get("month", now.month, type=int)
    year = request.args.get("year", now.year, type=int)
    if month < 1 or month > 12:
        month = now.month
    if year < 2000 or year > 2100:
        year = now.year

    month_prefix = f"{year}-{month:02d}"
    histories = (
        StockHistory.query
        .join(Product, StockHistory.product_id == Product.id)
        .filter(StockHistory.created_at.like(f"{month_prefix}%"))
        .order_by(StockHistory.created_at.asc())
        .all()
    )

    report_data = []
    total_added = 0
    total_removed = 0
    for h in histories:
        product = db.session.get(Product, h.product_id)
        product_title = product.title if product else "Produit supprimé"
        entry = {
            "date": h.created_at,
            "product": product_title,
            "operation": h.change_type,
            "quantity": h.quantity,
            "stock_before": h.stock_before,
            "stock_after": h.stock_after,
        }
        report_data.append(entry)
        if h.change_type == "ADD":
            total_added += h.quantity
        else:
            total_removed += h.quantity

    all_products = Product.query.order_by(Product.title.asc()).all()
    stock_status = []
    for p in all_products:
        qty = p.stock_quantity
        if qty == 0:
            status_label = "FINISHED"
            status_class = "out-of-stock"
        elif qty <= 5:
            status_label = "LOW STOCK"
            status_class = "low-stock"
        else:
            status_label = "IN STOCK"
            status_class = "in-stock"
        stock_status.append({
            "title": p.title,
            "photo": p.photo,
            "stock_quantity": qty,
            "status": status_label,
            "status_class": status_class,
        })
    total_products = len(all_products)
    total_stock_units = sum(p.stock_quantity for p in all_products)
    in_stock_count = sum(1 for p in all_products if p.stock_quantity > 5)
    low_stock_count = sum(1 for p in all_products if 0 < p.stock_quantity <= 5)
    finished_count = sum(1 for p in all_products if p.stock_quantity == 0)

    return render_template(
        "stock_report.html", report_data=report_data,
        total_added=total_added, total_removed=total_removed,
        selected_month=month, selected_year=year,
        stock_status=stock_status,
        total_products=total_products,
        total_stock_units=total_stock_units,
        in_stock_count=in_stock_count,
        low_stock_count=low_stock_count,
        finished_count=finished_count,
    )


@stock_bp.route('/report/pdf', methods=['POST'])
@login_required
def stock_report_pdf():
    data = request.get_json(silent=True) or {}
    month = int(data.get("month", datetime.now().month))
    year = int(data.get("year", datetime.now().year))
    month_prefix = f"{year}-{month:02d}"

    histories = (
        StockHistory.query
        .join(Product, StockHistory.product_id == Product.id)
        .filter(StockHistory.created_at.like(f"{month_prefix}%"))
        .order_by(StockHistory.created_at.asc())
        .all()
    )

    all_products = Product.query.order_by(Product.title.asc()).all()

    class StockReportPDF(FPDF):
        def header(self):
            self.set_y(14)
            logo_path = os.path.join(
                os.path.dirname(os.path.dirname(__file__)), "static", "logo.png"
            )
            if os.path.exists(logo_path):
                self.image(logo_path, x=self.l_margin, y=8, w=42)
            self.set_font("Helvetica", "B", 18)
            self.set_text_color(*GOLD)
            self.cell(0, 8, "COCINA ESPAÑOLA", align="R", new_x="LMARGIN", new_y="NEXT")
            self.set_font("Helvetica", "I", 13)
            self.cell(0, 6, "Art MDF", align="R", new_x="LMARGIN", new_y="NEXT")
            self.set_font("Helvetica", "", 8)
            self.set_text_color(*MID_GRAY)
            self.cell(0, 4, "Rue Tange Center Al hirafiyin N 25", align="R", new_x="LMARGIN", new_y="NEXT")
            self.cell(0, 4, "Imzouren AL Hoceima", align="R", new_x="LMARGIN", new_y="NEXT")
            self.ln(3)
            self.set_draw_color(*GOLD)
            self.set_line_width(0.5)
            y = self.get_y()
            self.line(self.l_margin, y, self.w - self.r_margin, y)
            self.ln(5)

        def footer(self):
            self.set_y(-22)
            self.set_draw_color(200, 200, 200)
            self.set_line_width(0.3)
            self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
            self.ln(3)
            self.set_font("Helvetica", "", 7)
            self.set_text_color(*GRAY_TEXT)
            pw = self.w - self.l_margin - self.r_margin
            cw = pw / 3
            self.set_x(self.l_margin)
            self.cell(cw, 4, "Rue Tange Center Al hirafiyin N 25", align="C")
            self.cell(cw, 4, "Tel: +212 6 71 68 75 98", align="C")
            self.cell(cw, 4, "Instagram: cocinaespanola", align="C")
            self.ln()
            self.set_x(self.l_margin)
            self.cell(cw, 4, "Imzouren AL Hoceima", align="C")
            self.cell(cw, 4, "", align="C")
            self.set_text_color(*DARK)
            self.cell(cw, 4, f"{self.page_no()}/{{nb}}", align="R")

    pdf = StockReportPDF()
    pdf.set_auto_page_break(auto=True, margin=28)
    pdf.set_margins(8, 10, 8)
    pdf.alias_nb_pages()
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*GOLD)
    pdf.cell(0, 10, "Rapport Mensuel de Stock", align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(*DARK)
    months_names = [
        "", "Janvier", "Fevrier", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Aout", "Septembre", "Octobre", "Novembre", "Decembre",
    ]
    gen_date = datetime.now().strftime("%d/%m/%Y %H:%M")
    pdf.cell(0, 6, f"Periode: {months_names[month]} {year}    |    Genere le: {gen_date}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(*DARK)
    pdf.cell(0, 8, "Section 1: Mouvements de Stock", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)

    if histories:
        col_w = [35, 50, 26, 22, 30, 30]
        headers = ["Date", "Produit", "Operation", "Qte", "Avant", "Apres"]
        row_h = 9

        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(*GRAY_BG)
        pdf.set_text_color(*DARK)
        pdf.set_draw_color(0, 0, 0)
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 10, h, border=1, align="C", fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", "", 9)
        total_added = 0
        total_removed = 0
        for idx, entry in enumerate(histories):
            product = db.session.get(Product, entry.product_id)
            pname = product.title if product else "-"
            date_str = entry.created_at[:10] if entry.created_at else ""
            op_text = "+" if entry.change_type == "ADD" else "-"

            if idx % 2 == 0:
                pdf.set_fill_color(255, 255, 255)
            else:
                pdf.set_fill_color(*GRAY_BG)

            pdf.set_text_color(*DARK)
            pdf.cell(col_w[0], row_h, date_str, border=1, fill=True)
            pdf.cell(col_w[1], row_h, pname[:28], border=1, fill=True)
            pdf.cell(col_w[2], row_h, op_text, border=1, align="C", fill=True)
            pdf.cell(col_w[3], row_h, str(entry.quantity), border=1, align="C", fill=True)
            pdf.cell(col_w[4], row_h, str(entry.stock_before), border=1, align="C", fill=True)
            pdf.cell(col_w[5], row_h, str(entry.stock_after), border=1, align="C", fill=True)
            pdf.ln()
            if entry.change_type == "ADD":
                total_added += entry.quantity
            else:
                total_removed += entry.quantity

        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(39, 174, 96)
        pdf.cell(0, 8, f"Total Ajoutes: {total_added}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(231, 76, 60)
        pdf.cell(0, 8, f"Total Retires: {total_removed}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(*DARK)
    else:
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*MID_GRAY)
        pdf.cell(0, 8, "Aucune operation de stock pour ce mois.", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(*DARK)

    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*GOLD)
    pdf.cell(0, 10, "Stock par Produit", align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(*DARK)
    pdf.cell(0, 6, f"Genere le: {gen_date}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    table_width = pdf.w - pdf.l_margin - pdf.r_margin
    product_col_w = table_width * 0.65
    stock_col_w = table_width * 0.20
    status_col_w = table_width * 0.15
    stock_col_w2 = [product_col_w, stock_col_w, status_col_w]
    stock_headers = ["Product", "Current Stock", "Status"]
    row_h = 12

    def draw_stock_table_header():
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(*GRAY_BG)
        pdf.set_text_color(*DARK)
        pdf.set_draw_color(0, 0, 0)
        alignments = ["L", "C", "C"]
        for i, h in enumerate(stock_headers):
            pdf.cell(stock_col_w2[i], 10, h, border=1, align=alignments[i], fill=True)
        pdf.ln()

    draw_stock_table_header()

    pdf.set_font("Helvetica", "", 10)
    last_page = pdf.page_no()
    for idx, product in enumerate(all_products):
        if pdf.page_no() != last_page:
            draw_stock_table_header()
            last_page = pdf.page_no()

        qty = product.stock_quantity
        if qty == 0:
            status = "FINISHED"
        elif qty <= 5:
            status = "LOW STOCK"
        else:
            status = "IN STOCK"

        x_start = pdf.get_x()
        y_start = pdf.get_y()

        if y_start + row_h > pdf.h - 28:
            pdf.add_page()
            draw_stock_table_header()
            last_page = pdf.page_no()
            x_start = pdf.get_x()
            y_start = pdf.get_y()

        pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(*DARK)
        pdf.set_draw_color(0, 0, 0)

        pdf.cell(product_col_w, row_h, product.title[:40], border=1, align="L", fill=True)

        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(stock_col_w, row_h, str(qty), border=1, align="C", fill=True)

        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(status_col_w, row_h, status, border=1, align="C", fill=True)

        pdf.ln()
        pdf.set_font("Helvetica", "", 10)
    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    name = f"rapport_stock_{year}_{month:02d}.pdf"
    return send_file(buf, as_attachment=True, download_name=name, mimetype="application/pdf")


@stock_bp.route('/report/excel', methods=['POST'])
@login_required
def stock_report_excel():
    data = request.get_json(silent=True) or {}
    month = int(data.get("month", datetime.now().month))
    year = int(data.get("year", datetime.now().year))
    month_prefix = f"{year}-{month:02d}"

    histories = (
        StockHistory.query
        .join(Product, StockHistory.product_id == Product.id)
        .filter(StockHistory.created_at.like(f"{month_prefix}%"))
        .order_by(StockHistory.created_at.asc())
        .all()
    )

    all_products = Product.query.order_by(Product.title.asc()).all()
    total_products = len(all_products)
    total_stock_units = sum(p.stock_quantity for p in all_products)
    in_stock_count = sum(1 for p in all_products if p.stock_quantity > 5)
    low_stock_count = sum(1 for p in all_products if 0 < p.stock_quantity <= 5)
    finished_count = sum(1 for p in all_products if p.stock_quantity == 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Stock"

    GOLD_HEX = "C9A84C"
    GRAY_HEX = "F0F0F0"
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    gray_fill = PatternFill(start_color=GRAY_HEX, end_color=GRAY_HEX, fill_type="solid")
    red_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    green_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")

    months_names = [
        "", "Janvier", "Fevrier", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Aout", "Septembre", "Octobre", "Novembre", "Decembre",
    ]
    gen_date = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws.merge_cells('A1:F1')
    ws['A1'] = "Rapport Mensuel de Stock"
    ws['A1'].font = Font(name="Arial", bold=True, size=16, color=GOLD_HEX)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Periode: {months_names[month]} {year}    |    Genere le: {gen_date}"
    ws['A2'].font = Font(name="Arial", size=10, color="555555")
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A4:F4')
    ws['A4'] = "Section 1: Mouvements de Stock"
    ws['A4'].font = Font(name="Arial", bold=True, size=12, color="333333")

    headers = ["Date", "Produit", "Operation", "Quantite", "Stock Avant", "Stock Apres"]
    col_widths = [16, 30, 12, 10, 12, 12]
    header_row = 6
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="333333")
        cell.fill = gray_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = col_widths[col - 1]

    total_added = 0
    total_removed = 0
    for i, entry in enumerate(histories):
        r = header_row + 1 + i
        product = db.session.get(Product, entry.product_id)
        pname = product.title if product else "-"
        date_str = entry.created_at[:10] if entry.created_at else ""
        op_text = "Ajout" if entry.change_type == "ADD" else "Retrait"

        ws.cell(row=r, column=1, value=date_str).border = thin_border
        ws.cell(row=r, column=2, value=pname).border = thin_border
        ws.cell(row=r, column=3, value=op_text).border = thin_border
        ws.cell(row=r, column=4, value=entry.quantity).border = thin_border
        ws.cell(row=r, column=5, value=entry.stock_before).border = thin_border
        ws.cell(row=r, column=6, value=entry.stock_after).border = thin_border
        for c in range(1, 7):
            ws.cell(row=r, column=c).font = Font(name="Arial", size=10)
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
        if entry.change_type == "ADD":
            total_added += entry.quantity
        else:
            total_removed += entry.quantity

    summary_row = header_row + 1 + len(histories) + 1
    ws.cell(row=summary_row, column=2, value="Total Ajoutes").font = Font(name="Arial", bold=True, size=11, color="27ae60")
    ws.cell(row=summary_row, column=3, value=total_added).font = Font(name="Arial", bold=True, size=11, color="27ae60")
    ws.cell(row=summary_row + 1, column=2, value="Total Retires").font = Font(name="Arial", bold=True, size=11, color="e74c3c")
    ws.cell(row=summary_row + 1, column=3, value=total_removed).font = Font(name="Arial", bold=True, size=11, color="e74c3c")

    sec2_start = summary_row + 4
    ws.merge_cells(f'A{sec2_start}:E{sec2_start}')
    ws[f'A{sec2_start}'] = "Section 2: Etat du Stock Actuel"
    ws[f'A{sec2_start}'].font = Font(name="Arial", bold=True, size=12, color="333333")

    stat_labels = ["Produits", "En Stock", "Stock Bas", "Termine", "Unites Total"]
    stat_vals = [total_products, in_stock_count, low_stock_count, finished_count, total_stock_units]
    stat_row = sec2_start + 2
    for i, (lbl, val) in enumerate(zip(stat_labels, stat_vals)):
        c = i + 1
        lbl_cell = ws.cell(row=stat_row, column=c, value=lbl)
        lbl_cell.font = Font(name="Arial", size=9, color="555555")
        lbl_cell.fill = gray_fill
        lbl_cell.alignment = Alignment(horizontal='center')
        lbl_cell.border = thin_border
        val_cell = ws.cell(row=stat_row + 1, column=c, value=val)
        val_cell.font = Font(name="Arial", bold=True, size=11, color="333333")
        val_cell.alignment = Alignment(horizontal='center')
        val_cell.border = thin_border

    s2_headers = ["#", "Produit", "Stock Actuel", "Statut"]
    s2_header_row = stat_row + 4
    for col, h in enumerate(s2_headers, 1):
        cell = ws.cell(row=s2_header_row, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="333333")
        cell.fill = gray_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for idx, item in enumerate(stock_status_from_db(all_products), 1):
        r = s2_header_row + idx
        qty = item["stock_quantity"]
        status = item["status"]

        ws.cell(row=r, column=1, value=idx).border = thin_border
        ws.cell(row=r, column=2, value=item["title"]).border = thin_border
        ws.cell(row=r, column=3, value=qty).border = thin_border
        ws.cell(row=r, column=4, value=status).border = thin_border

        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='left')

        if status in ("FINISHED", "LOW STOCK"):
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = red_fill
                ws.cell(row=r, column=c).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        else:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = green_fill
                ws.cell(row=r, column=c).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    name = f"rapport_stock_{year}_{month:02d}.xlsx"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(tmp.name, as_attachment=True, download_name=name)


@stock_bp.route('/status')
@login_required
def stock_status_report():
    data = _build_stock_status_data()
    return render_template("stock_status.html", **data)


@stock_bp.route('/api/data')
@login_required
def api_stock_data():
    data = _build_stock_status_data()
    return jsonify(data)


@stock_bp.route('/status/pdf', methods=['POST'])
@login_required
def stock_status_pdf():
    data = _build_stock_status_data()

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    GOLD_RGB = (201, 168, 76)
    DARK_RGB = (51, 51, 51)
    GRAY_RGB = (240, 240, 240)
    RED_RGB = (231, 76, 60)
    ORANGE_RGB = (230, 126, 34)

    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(*GOLD_RGB)
    pdf.cell(0, 12, "Etat du Stock", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(*DARK_RGB)
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    pdf.cell(0, 8, f"Genere le: {now_str}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, f"Total Produits: {data['total_products']}   |   Total Stock: {data['total_stock']}   |   Stock Bas: {data['low_stock']}   |   Rupture: {data['out_of_stock']}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    col_widths = [14, 55, 30, 40, 35]
    headers = ["#", "Produit", "Stock", "Statut", "Maj le"]

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(*GRAY_RGB)
    pdf.set_text_color(*DARK_RGB)
    pdf.set_draw_color(200, 200, 200)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 8, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    for idx, row in enumerate(data["rows"], 1):
        qty = row["stock_quantity"]
        status = row["status"]
        if status == "OUT OF STOCK":
            pdf.set_fill_color(*RED_RGB)
            pdf.set_text_color(255, 255, 255)
        elif status == "LOW STOCK":
            pdf.set_fill_color(*ORANGE_RGB)
            pdf.set_text_color(255, 255, 255)
        else:
            pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(*DARK_RGB)

        use_fill = status != "IN STOCK"
        pdf.cell(col_widths[0], 7, str(idx), border=1, align="C", fill=use_fill)
        pdf.cell(col_widths[1], 7, row["title"][:28], border=1, fill=use_fill)
        pdf.cell(col_widths[2], 7, str(qty), border=1, align="C", fill=use_fill)
        pdf.cell(col_widths[3], 7, status, border=1, align="C", fill=use_fill)
        pdf.cell(col_widths[4], 7, row["updated_at"][:16], border=1, fill=use_fill)
        pdf.ln()

    pdf.set_text_color(*DARK_RGB)
    pdf.ln(6)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, f"Total Produits: {data['total_products']}   |   Total Unites en Stock: {data['total_stock']}", align="C")

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    name = "etat_stock.pdf"
    return send_file(buf, as_attachment=True, download_name=name, mimetype="application/pdf")


@stock_bp.route('/status/excel', methods=['POST'])
@login_required
def stock_status_excel():
    data = _build_stock_status_data()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Etat du Stock"

    GOLD_HEX = "C9A84C"
    GRAY_HEX = "F0F0F0"
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    gray_fill = PatternFill(start_color=GRAY_HEX, end_color=GRAY_HEX, fill_type="solid")
    red_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    orange_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")

    ws.merge_cells('A1:E1')
    ws['A1'] = "Etat du Stock"
    ws['A1'].font = Font(name="Arial", bold=True, size=16, color=GOLD_HEX)
    ws['A1'].alignment = Alignment(horizontal='center')

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Genere le: {now_str}"
    ws['A2'].font = Font(name="Arial", size=10, color="555555")
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:E3')
    ws['A3'] = f"Total: {data['total_products']}  |  Stock: {data['total_stock']}  |  Bas: {data['low_stock']}  |  Rupture: {data['out_of_stock']}"
    ws['A3'].font = Font(name="Arial", size=10, color="333333")
    ws['A3'].alignment = Alignment(horizontal='center')

    headers = ["#", "Produit", "Stock", "Statut", "Derniere Maj"]
    col_widths = [6, 35, 10, 16, 20]
    header_row = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="333333")
        cell.fill = gray_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = col_widths[col - 1]

    for idx, row in enumerate(data["rows"], 1):
        r = header_row + idx
        qty = row["stock_quantity"]
        status = row["status"]

        ws.cell(row=r, column=1, value=idx).border = thin_border
        ws.cell(row=r, column=2, value=row["title"]).border = thin_border
        ws.cell(row=r, column=3, value=qty).border = thin_border
        status_cell = ws.cell(row=r, column=4, value=status)
        status_cell.border = thin_border
        ws.cell(row=r, column=5, value=row["updated_at"][:16]).border = thin_border

        for c in range(1, 6):
            ws.cell(row=r, column=c).font = Font(name="Arial", size=10)
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')

        ws.cell(row=r, column=2).alignment = Alignment(horizontal='left')

        if status == "OUT OF STOCK":
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = red_fill
                ws.cell(row=r, column=c).font = Font(name="Arial", size=10, color="FFFFFF")
            status_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        elif status == "LOW STOCK":
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = orange_fill
                ws.cell(row=r, column=c).font = Font(name="Arial", size=10, color="FFFFFF")
            status_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        else:
            status_cell.font = Font(name="Arial", size=10, bold=True, color="27ae60")

    name = "etat_stock.xlsx"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(tmp.name, as_attachment=True, download_name=name)
