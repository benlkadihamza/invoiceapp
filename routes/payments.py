import csv
import tempfile
from io import BytesIO, StringIO
from flask import Blueprint, render_template, request, jsonify, send_file, redirect, url_for, flash
from flask_login import login_required
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from models import db, Invoice, InvoicePayment

payments_bp = Blueprint('payments', __name__, url_prefix='/payments')

GOLD = (201, 168, 76)
DARK = (51, 51, 51)
GRAY_BG = (240, 240, 240)


def sync_invoice_payments():
    invoices = Invoice.query.all()
    for inv in invoices:
        pm = InvoicePayment.query.filter_by(invoice_id=inv.id).first()
        if not pm:
            pm = InvoicePayment(
                invoice_id=inv.id,
                customer_name=inv.client_name,
                invoice_number=inv.invoice_num,
                invoice_total=inv.net_total,
                remaining_amount=inv.net_total,
                status="Pending"
            )
            db.session.add(pm)
        else:
            pm.customer_name = inv.client_name
            pm.invoice_number = inv.invoice_num
            pm.invoice_total = inv.net_total
            pm.recalculate()
    db.session.commit()


@payments_bp.route('/')
@login_required
def list_payments():
    sync_invoice_payments()
    payments = InvoicePayment.query.order_by(InvoicePayment.invoice_id.desc(), InvoicePayment.id.desc()).all()

    pending_count = sum(1 for p in payments if p.status == 'Pending')
    partial_count = sum(1 for p in payments if p.status == 'Partial')
    paid_count = sum(1 for p in payments if p.status == 'Paid')
    total_remaining = sum(p.remaining_amount for p in payments)
    total_collected = sum(p.total_paid for p in payments)

    return render_template(
        'payments.html',
        payments=payments,
        pending_count=pending_count,
        partial_count=partial_count,
        paid_count=paid_count,
        total_remaining=total_remaining,
        total_collected=total_collected
    )


@payments_bp.route('/<int:payment_id>')
@login_required
def payment_detail(payment_id):
    pm = db.session.get(InvoicePayment, payment_id)
    if not pm:
        flash("Fiche de paiement introuvable.", "danger")
        return redirect(url_for('payments.list_payments'))
    
    invoice = db.session.get(Invoice, pm.invoice_id)
    return render_template('payment_detail.html', payment=pm, invoice=invoice)


from idempotency import idempotent_route


@payments_bp.route('/<int:payment_id>/update', methods=['POST'])
@login_required
@idempotent_route()
def update_payment(payment_id):
    pm = db.session.get(InvoicePayment, payment_id)
    is_ajax = request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if not pm:
        if is_ajax:
            return jsonify({"error": "Fiche de paiement introuvable."}), 404
        flash("Fiche de paiement introuvable.", "danger")
        return redirect(url_for('payments.list_payments'))

    data = request.get_json(silent=True) if request.is_json else request.form

    def parse_float(val):
        try:
            return float(val) if val is not None and str(val).strip() != "" else 0.0
        except ValueError:
            return 0.0

    p1 = parse_float(data.get('payment1_amount'))
    p2 = parse_float(data.get('payment2_amount'))
    p3 = parse_float(data.get('payment3_amount'))
    p4 = parse_float(data.get('payment4_amount'))

    # Validation
    if p1 < 0 or p2 < 0 or p3 < 0 or p4 < 0:
        msg = "Les montants des paiements ne peuvent pas être négatifs."
        if is_ajax:
            return jsonify({"error": msg}), 400
        flash(msg, "danger")
        return redirect(url_for('payments.payment_detail', payment_id=pm.id))

    total_paid = round(p1 + p2 + p3 + p4, 2)
    inv_total = round(pm.invoice_total, 2)
    if total_paid > inv_total:
        exceed_amount = round(total_paid - inv_total, 2)
        msg = f"Erreur : Le montant du paiement dépasse le reste à payer de la facture ! (Total saisi : {total_paid:,.2f} DH | Total facture : {inv_total:,.2f} DH | Dépassement : {exceed_amount:,.2f} DH)"
        if is_ajax:
            return jsonify({"error": msg}), 400
        flash(msg, "danger")
        return redirect(url_for('payments.payment_detail', payment_id=pm.id))

    try:
        pm.payment1_amount = p1
        pm.payment1_date = data.get('payment1_date', '') or ''
        pm.payment1_notes = data.get('payment1_notes', '') or ''

        pm.payment2_amount = p2
        pm.payment2_date = data.get('payment2_date', '') or ''
        pm.payment2_notes = data.get('payment2_notes', '') or ''

        pm.payment3_amount = p3
        pm.payment3_date = data.get('payment3_date', '') or ''
        pm.payment3_notes = data.get('payment3_notes', '') or ''

        pm.payment4_amount = p4
        pm.payment4_date = data.get('payment4_date', '') or ''
        pm.payment4_notes = data.get('payment4_notes', '') or ''

        pm.recalculate()
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        msg = f"Erreur lors de la mise à jour : {str(e)}"
        if is_ajax:
            return jsonify({"error": msg}), 500
        flash(msg, "danger")
        return redirect(url_for('payments.payment_detail', payment_id=pm.id))

    is_complete = (pm.status == 'Paid') or (pm.remaining_amount <= 0)

    if is_ajax:
        return jsonify({
            "success": True,
            "message": "Paiement complété avec succès ! Statut: PAYÉ" if is_complete else "Paiement mis à jour avec succès !",
            "status": pm.status,
            "remaining_amount": pm.remaining_amount,
            "total_paid": pm.total_paid,
            "is_complete": is_complete,
            "redirect": url_for('payments.list_payments') if is_complete else None
        })

    if is_complete:
        flash(f"Paiement de la facture N° {pm.invoice_number or f'#{pm.invoice_id}'} complété avec succès ! Statut: PAYÉ", "success")
        return redirect(url_for('payments.list_payments'))

    flash("Paiement mis à jour avec succès !", "success")
    return redirect(url_for('payments.payment_detail', payment_id=pm.id))


@payments_bp.route('/export/excel')
@login_required
def export_excel():
    sync_invoice_payments()
    payments = InvoicePayment.query.order_by(InvoicePayment.invoice_id.desc(), InvoicePayment.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gestion des Paiements"

    GOLD_HEX = "C9A84C"
    GRAY_HEX = "F0F0F0"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color=GRAY_HEX, end_color=GRAY_HEX, fill_type="solid")

    headers = [
        "N° Facture", "Client", "Total Facture (DH)",
        "Paiement 1 (DH)", "Paiement 2 (DH)", "Paiement 3 (DH)", "Paiement 4 (DH)",
        "Reste à Payer (DH)", "Statut", "Date de mise à jour"
    ]

    ws.merge_cells('A1:J1')
    ws['A1'] = "RAPPORT DE GESTION DES PAIEMENTS"
    ws['A1'].font = Font(name="Arial", bold=True, size=16, color=GOLD_HEX)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 30

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="333333")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for i, pm in enumerate(payments, start=4):
        ws.cell(row=i, column=1, value=pm.invoice_number or f"#{pm.invoice_id}").border = thin_border
        ws.cell(row=i, column=2, value=pm.customer_name).border = thin_border
        ws.cell(row=i, column=3, value=pm.invoice_total).border = thin_border
        ws.cell(row=i, column=4, value=pm.payment1_amount).border = thin_border
        ws.cell(row=i, column=5, value=pm.payment2_amount).border = thin_border
        ws.cell(row=i, column=6, value=pm.payment3_amount).border = thin_border
        ws.cell(row=i, column=7, value=pm.payment4_amount).border = thin_border
        ws.cell(row=i, column=8, value=pm.remaining_amount).border = thin_border
        ws.cell(row=i, column=9, value="PAYÉ" if pm.status == "Paid" else ("PARTIEL" if pm.status == "Partial" else "EN ATTENTE")).border = thin_border
        ws.cell(row=i, column=10, value=pm.updated_at or pm.created_at).border = thin_border

        for c in range(1, 11):
            ws.cell(row=i, column=c).font = Font(name="Arial", size=9)
            if c in [3, 4, 5, 6, 7, 8]:
                ws.cell(row=i, column=c).alignment = Alignment(horizontal='right')
            else:
                ws.cell(row=i, column=c).alignment = Alignment(horizontal='center')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(tmp.name, as_attachment=True, download_name="gestion_des_paiements.xlsx")


@payments_bp.route('/export/csv')
@login_required
def export_csv():
    sync_invoice_payments()
    payments = InvoicePayment.query.order_by(InvoicePayment.invoice_id.desc(), InvoicePayment.id.desc()).all()

    si = StringIO()
    cw = csv.writer(si)
    cw.writerow([
        "N° Facture", "Client", "Total Facture (DH)",
        "Paiement 1 (DH)", "Paiement 2 (DH)", "Paiement 3 (DH)", "Paiement 4 (DH)",
        "Reste à Payer (DH)", "Statut", "Date de mise à jour"
    ])

    for pm in payments:
        cw.writerow([
            pm.invoice_number or f"#{pm.invoice_id}",
            pm.customer_name,
            f"{pm.invoice_total:.2f}",
            f"{pm.payment1_amount:.2f}",
            f"{pm.payment2_amount:.2f}",
            f"{pm.payment3_amount:.2f}",
            f"{pm.payment4_amount:.2f}",
            f"{pm.remaining_amount:.2f}",
            "PAYÉ" if pm.status == "Paid" else ("PARTIEL" if pm.status == "Partial" else "EN ATTENTE"),
            pm.updated_at or pm.created_at
        ])

    buf = BytesIO()
    buf.write(si.getvalue().encode('utf-8-sig'))
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="gestion_des_paiements.csv", mimetype="text/csv")


class PaymentsPDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 16)
        self.set_text_color(*GOLD)
        self.cell(0, 8, "GESTION DES PAIEMENTS", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 10)
        self.set_text_color(*DARK)
        self.cell(0, 6, "Rapport recapitulatif des reglements par tranche", align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(6)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "", 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")


@payments_bp.route('/export/pdf')
@login_required
def export_pdf():
    sync_invoice_payments()
    payments = InvoicePayment.query.order_by(InvoicePayment.invoice_id.desc(), InvoicePayment.id.desc()).all()

    pdf = PaymentsPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    col_w = [25, 45, 28, 25, 25, 25, 25, 28, 25, 22]
    headers = ["N° Fact", "Client", "Total DH", "Paiem 1", "Paiem 2", "Paiem 3", "Paiem 4", "Reste DH", "Statut", "Date"]

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(*GRAY_BG)
    pdf.set_text_color(*DARK)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    for pm in payments:
        status_str = "PAYÉ" if pm.status == "Paid" else ("PARTIEL" if pm.status == "Partial" else "EN ATTENTE")
        pdf.cell(col_w[0], 6, str(pm.invoice_number or f"#{pm.invoice_id}")[:12], border=1, align="C")
        pdf.cell(col_w[1], 6, str(pm.customer_name)[:22], border=1, align="L")
        pdf.cell(col_w[2], 6, f"{pm.invoice_total:,.2f}".replace(',', ' '), border=1, align="R")
        pdf.cell(col_w[3], 6, f"{pm.payment1_amount:,.2f}".replace(',', ' '), border=1, align="R")
        pdf.cell(col_w[4], 6, f"{pm.payment2_amount:,.2f}".replace(',', ' '), border=1, align="R")
        pdf.cell(col_w[5], 6, f"{pm.payment3_amount:,.2f}".replace(',', ' '), border=1, align="R")
        pdf.cell(col_w[6], 6, f"{pm.payment4_amount:,.2f}".replace(',', ' '), border=1, align="R")
        pdf.cell(col_w[7], 6, f"{pm.remaining_amount:,.2f}".replace(',', ' '), border=1, align="R")
        pdf.cell(col_w[8], 6, status_str, border=1, align="C")
        date_str = (pm.updated_at or pm.created_at or "").split(" ")[0]
        pdf.cell(col_w[9], 6, date_str, border=1, align="C")
        pdf.ln()

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="gestion_des_paiements.pdf", mimetype="application/pdf")
