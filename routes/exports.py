from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from models import db, Transaction, Person, PaymentMethod
from sqlalchemy import func, extract
from datetime import date, datetime
from io import StringIO, BytesIO
import csv
import os

exports_bp = Blueprint('exports', __name__)


def _get_filtered_transactions():
    query = Transaction.query
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    person_id = request.args.get('person_id', 0, type=int)

    if date_from:
        try:
            query = query.filter(Transaction.date >= date.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Transaction.date <= date.fromisoformat(date_to))
        except ValueError:
            pass
    if person_id:
        query = query.filter(Transaction.person_id == person_id)

    return query.order_by(Transaction.date.asc(), Transaction.id.asc()).all()


@exports_bp.route('/export/csv')
@login_required
def export_csv():
    transactions = _get_filtered_transactions()

    si = StringIO()
    writer = csv.writer(si, delimiter=';', lineterminator='\n')
    writer.writerow(['Date', 'Personne', 'Description', 'Revenu (DH)', 'Dépense (DH)', 'Net (DH)', 'Mode de paiement', 'Notes'])

    balance = 0.0
    for t in transactions:
        balance += t.income - t.expense
        writer.writerow([
            t.formatted_date,
            t.person.name if t.person else '',
            t.description,
            f'{t.income:.2f}', f'{t.expense:.2f}',
            f'{t.income - t.expense:.2f}',
            t.payment_method.name if t.payment_method else '',
            t.notes or ''
        ])

    from flask import Response
    output = '\ufeff' + si.getvalue()
    return Response(
        output.encode('utf-8-sig'),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment;filename=transactions.csv'}
    )


@exports_bp.route('/export/excel')
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    transactions = _get_filtered_transactions()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2D3436', end_color='2D3436', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Date', 'Personne', 'Description', 'Revenu (DH)', 'Dépense (DH)', 'Net (DH)', 'Mode de paiement', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    balance = 0.0
    for row_idx, t in enumerate(transactions, 2):
        balance += t.income - t.expense
        values = [
            t.formatted_date,
            t.person.name if t.person else '',
            t.description,
            t.income, t.expense,
            t.income - t.expense,
            t.payment_method.name if t.payment_method else '',
            t.notes or ''
        ]
        income_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        expense_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
        fill = income_fill if t.income > 0 else expense_fill

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col in (4, 5, 6):
                cell.number_format = '#,##0.00'
            if col == 4:
                cell.fill = fill

    total_row = len(transactions) + 2
    ws.cell(row=total_row, column=3, value='TOTAUX').font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(t.income for t in transactions)).font = Font(bold=True)
    ws.cell(row=total_row, column=4).number_format = '#,##0.00'
    ws.cell(row=total_row, column=5, value=sum(t.expense for t in transactions)).font = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = '#,##0.00'
    ws.cell(row=total_row, column=6, value=sum(t.income - t.expense for t in transactions)).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = '#,##0.00'

    col_widths = [12, 15, 30, 15, 15, 15, 15, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='transactions.xlsx')


@exports_bp.route('/export/pdf')
@login_required
def export_pdf():
    return redirect(url_for('reports.full_pdf'))
