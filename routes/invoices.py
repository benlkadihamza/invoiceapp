import os
import re
import tempfile
from io import BytesIO
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from models import db, Invoice

invoices_bp = Blueprint('invoices', __name__, url_prefix='/invoices')

COMPANY_NAME = "Votre Société"
COMPANY_ADDRESS = "Rue Tange Center Al hirafiyin N 25 Imzouren Al Hoceima"
COMPANY_PHONE = ""
COMPANY_EMAIL = ""

GOLD = (201, 168, 76)
DARK = (51, 51, 51)
GRAY_BG = (240, 240, 240)
GRAY_TEXT = (136, 136, 136)
MID_GRAY = (85, 85, 85)


def safe_filename(name):
    name = re.sub(r'[<>:"/\\|?*]', '_', name).strip()
    if not name:
        name = "facture"
    return name


def save_invoice_data(data):
    invoice_num = data.get("invoice_num", "")
    date_val = data.get("date", "")
    client_name = data.get("client_name", "")
    client_address = data.get("client_address", "")
    total = data.get("total", 0.0)
    remise = data.get("remise", 0.0)
    payer = data.get("payer", 0.0)
    net_total = data.get("net_total", 0.0)
    items = data.get("items", [])
    show_facture_num = bool(data.get("show_facture_num", False))

    existing_id = data.get("id")

    if existing_id:
        invoice = db.session.get(Invoice, existing_id)
        if invoice:
            invoice.invoice_num = invoice_num
            invoice.date = date_val
            invoice.client_name = client_name
            invoice.client_address = client_address
            invoice.total = total
            invoice.remise = remise
            invoice.payer = payer
            invoice.net_total = net_total
            invoice.items = items
            invoice.show_facture_num = show_facture_num
        invoice_id = existing_id
    else:
        invoice = Invoice(
            invoice_num=invoice_num,
            date=date_val,
            client_name=client_name,
            client_address=client_address,
            total=total,
            remise=remise,
            payer=payer,
            net_total=net_total,
            items=items,
            show_facture_num=show_facture_num
        )
        db.session.add(invoice)
        db.session.flush()
        invoice_id = invoice.id

    db.session.commit()

    from models import InvoicePayment
    pm = InvoicePayment.query.filter_by(invoice_id=invoice_id).first()
    if not pm:
        pm = InvoicePayment(
            invoice_id=invoice_id,
            customer_name=client_name,
            invoice_number=invoice_num,
            invoice_total=net_total,
            remaining_amount=net_total,
            status="Pending"
        )
        db.session.add(pm)
    else:
        pm.customer_name = client_name
        pm.invoice_number = invoice_num
        pm.invoice_total = net_total
        pm.recalculate()

    db.session.commit()
    return invoice_id


def get_invoice_by_id(invoice_id):
    invoice = db.session.get(Invoice, invoice_id)
    if not invoice:
        return None
    return {
        "id": invoice.id,
        "invoice_num": invoice.invoice_num,
        "date": invoice.date,
        "client_name": invoice.client_name,
        "client_address": invoice.client_address,
        "total": invoice.total,
        "remise": invoice.remise,
        "payer": invoice.payer,
        "net_total": invoice.net_total,
        "items": invoice.items,
        "show_facture_num": invoice.show_facture_num
    }


class InvoicePDF(FPDF):
    def header(self):
        if self.page_no() > 1:
            return

        self.set_y(25)

        logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "logo.png")
        if os.path.exists(logo_path):
            self.image(logo_path, x=self.l_margin, y=14, w=48)

        self.set_font("Helvetica", "B", 20)
        self.set_text_color(*GOLD)
        self.cell(0, 8, "COCINA ESPAÑOLA", align="R", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "I", 14)
        self.cell(0, 6, "Art MDF", align="R", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 9)
        self.set_text_color(*MID_GRAY)
        self.cell(0, 5, "Rue Tange Center Al hirafiyin N 25", align="R", new_x="LMARGIN", new_y="NEXT")
        self.cell(0, 5, "Imzouren AL Hoceima", align="R", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "B", 9)
        self.set_text_color(*DARK)
        client_name = getattr(self, "_client_name", "")
        self.cell(0, 5, f"Client: {client_name}", align="R", new_x="LMARGIN", new_y="NEXT")
        self.ln(10)

        self.set_draw_color(*GOLD)
        self.set_line_width(0.6)
        self.set_font("Helvetica", "B", 18)
        self.set_text_color(*DARK)
        facture_num = getattr(self, "_facture_num", "")
        show_fn = getattr(self, "_show_facture_num", False)
        title = "FACTURE"
        if show_fn and facture_num:
            title += f" N\u00b0 {facture_num}"
        self.cell(0, 10, title, align="C", new_x="LMARGIN", new_y="NEXT")
        y = self.get_y()
        self.line(95, y, 115, y)
        self.ln(8)

        date_val = getattr(self, "_date", "")
        self.set_font("Helvetica", "", 10)
        self.set_text_color(*MID_GRAY)
        self.cell(0, 6, f"Date de facture: {date_val}", align="R", new_x="LMARGIN", new_y="NEXT")
        self.ln(6)

    def footer(self):
        self.set_y(-25)
        self.set_draw_color(200, 200, 200)
        self.set_line_width(0.3)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.ln(4)

        self.set_font("Helvetica", "", 7)
        self.set_text_color(*GRAY_TEXT)
        page_w = self.w - self.l_margin - self.r_margin
        col_w = page_w / 3
        self.set_x(self.l_margin)
        self.cell(col_w, 4, "Rue Tange Center Al hirafiyin N 25", align="C")
        self.cell(col_w, 4, "Tel: +212 6 71 68 75 98", align="C")
        self.cell(col_w, 4, "Instagram: cocinaespanola", align="C")
        self.ln()
        self.set_x(self.l_margin)
        self.cell(col_w, 4, "Imzouren AL Hoceima", align="C")
        self.cell(col_w, 4, "", align="C")
        self.set_text_color(*DARK)
        self.cell(col_w, 4, f"{self.page_no()}/{{nb}}", align="R")


@invoices_bp.route('/new')
@login_required
def create_invoice():
    return render_template('invoice_create.html')


@invoices_bp.route('/')
@login_required
def list_invoices():
    try:
        from models import InvoicePayment
        rows = Invoice.query.order_by(Invoice.id.desc()).all()
        invoices = []
        for inv in rows:
            pm = InvoicePayment.query.filter_by(invoice_id=inv.id).first()
            if not pm:
                pm = InvoicePayment(
                    invoice_id=inv.id,
                    customer_name=inv.client_name,
                    invoice_number=inv.invoice_num,
                    invoice_total=inv.net_total,
                    remaining_amount=inv.net_total,
                    status="Pending"
                )
                db.session.add(pm)
                db.session.commit()
            invoices.append({
                "id": inv.id,
                "invoice_num": inv.invoice_num,
                "date": inv.date,
                "client_name": inv.client_name,
                "net_total": inv.net_total,
                "created_at": inv.created_at,
                "payment_status": pm.status,
                "remaining_amount": pm.remaining_amount,
                "total_paid": pm.total_paid
            })
    except Exception:
        invoices = []
    return render_template('invoices.html', invoices=invoices)


from idempotency import idempotent_route


@invoices_bp.route('/save', methods=['POST'])
@login_required
@idempotent_route()
def save_invoice_route():
    data = request.get_json(silent=True) or {}
    try:
        is_update = bool(data.get("id"))
        invoice_id = save_invoice_data(data)
        if invoice_id is None:
            return jsonify({"error": "Facture non trouvée"}), 404
        
        from flask import flash, url_for
        client_str = f" ({data.get('client_name')})" if data.get('client_name') else ""
        if is_update:
            flash(f"La facture N° {data.get('invoice_num', '')}{client_str} a été mise à jour avec succès.", "success")
        else:
            flash(f"La facture N° {data.get('invoice_num', '')}{client_str} a été créée avec succès.", "success")

        return jsonify({
            "id": invoice_id,
            "is_update": is_update,
            "redirect": url_for('invoices.list_invoices')
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erreur lors de l'enregistrement: {str(e)}"}), 500


@invoices_bp.route('/<int:invoice_id>/json')
@login_required
def get_invoice_json(invoice_id):
    data = get_invoice_by_id(invoice_id)
    if not data:
        return jsonify({"error": "Facture non trouvée"}), 404
    return jsonify(data)


@invoices_bp.route('/generate_pdf', methods=['POST'])
@login_required
def generate_pdf():
    data = request.get_json(silent=True) or {}

    pdf = InvoicePDF()
    pdf._client_name = data.get("client_name", "")
    pdf._date = data.get("date", "")
    pdf._facture_num = data.get("invoice_num", "")
    pdf._show_facture_num = data.get("show_facture_num", False)
    pdf.set_auto_page_break(auto=True, margin=30)
    pdf.add_page()

    col_w = [72, 28, 44, 38]
    headers = ["Description", "Quantité", "Prix DH", "Montant DH"]

    def draw_table_header():
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(*GRAY_BG)
        pdf.set_text_color(*DARK)
        pdf.set_draw_color(0, 0, 0)
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 8, h, border=1, align="C", fill=True)
        pdf.ln()

    draw_table_header()

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*DARK)
    last_page = pdf.page_no()
    for item in data.get('items', []):
        if pdf.page_no() != last_page:
            draw_table_header()
            last_page = pdf.page_no()
        pdf.cell(col_w[0], 7, item.get('description', ''), border=1)
        pdf.cell(col_w[1], 7, str(item.get('quantity', 0)), border=1, align="C")
        pdf.cell(col_w[2], 7, f"{item.get('unit_price', 0):,.2f}".replace(',', ' '), border=1, align="C")
        pdf.cell(col_w[3], 7, f"{item.get('total', 0):,.2f}".replace(',', ' '), border=1, align="C")
        pdf.ln()

    total = data.get('total', 0.0)
    remise = data.get('remise', 0.0)
    payer = data.get('payer', 0.0)
    net_total = data.get('net_total', total)

    pdf.ln(4)
    total_x = pdf.l_margin + col_w[0] + col_w[1]

    def price_line(label, value, val_color=None, bold=False):
        pdf.set_x(total_x)
        pdf.set_fill_color(*GRAY_BG)
        pdf.set_draw_color(0, 0, 0)
        pdf.set_font("Helvetica", "B" if bold else "", 10)
        pdf.set_text_color(*(val_color if val_color else DARK))
        pdf.cell(col_w[2], 10, label, border=1, align="C", fill=True)
        pdf.cell(col_w[3], 10, value, border=1, align="C", fill=True)
        pdf.ln(10)

    if remise > 0 or payer > 0:
        price_line("Total en DH", f"{total:,.2f} DH".replace(',', ' '))
    if remise > 0:
        price_line("Remise en DH", f"-{remise:,.2f} DH".replace(',', ' '))
    if payer > 0:
        price_line("Payer en DH", f"-{payer:,.2f} DH".replace(',', ' '))
    price_line("Total a Payer en DH", f"{net_total:,.2f} DH".replace(',', ' '), val_color=GOLD, bold=True)

    name = f"{safe_filename(data.get('client_name', ''))}.pdf"
    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=name, mimetype="application/pdf")


@invoices_bp.route('/generate_excel', methods=['POST'])
@login_required
def generate_excel():
    data = request.get_json(silent=True) or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facture"

    GOLD_HEX = "C9A84C"
    GRAY_HEX = "F0F0F0"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color=GRAY_HEX, end_color=GRAY_HEX, fill_type="solid")

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    ws.merge_cells('A1:C1')
    ws['A1'] = "COCINA ESPAÑOLA"
    ws['A1'].font = Font(name="Arial", bold=True, size=20, color=GOLD_HEX)

    ws.merge_cells('A2:C2')
    ws['A2'] = "Art MDF"
    ws['A2'].font = Font(name="Arial", italic=True, size=14, color=GOLD_HEX)

    ws.merge_cells('D1:D2')
    ws['D1'] = "Rue Tange Center Al hirafiyin N 25"
    ws['D1'].font = Font(name="Arial", size=9, color="555555")
    ws['D1'].alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('D3:D3')
    ws['D3'] = "Imzouren AL Hoceima"
    ws['D3'].font = Font(name="Arial", size=9, color="555555")
    ws['D3'].alignment = Alignment(horizontal='right')

    ws['C4'] = "Client:"
    ws['C4'].font = Font(name="Arial", bold=True, size=10, color="333333")
    ws['C4'].alignment = Alignment(horizontal='right')

    ws['D4'] = data.get("client_name", "")
    ws['D4'].font = Font(name="Arial", bold=True, size=10, color="333333")
    ws['D4'].alignment = Alignment(horizontal='right')

    ws.merge_cells('A6:D6')
    title = "FACTURE"
    show_fn = data.get("show_facture_num", False)
    fn = data.get("invoice_num", "")
    if show_fn and fn:
        title += f" N\u00b0 {fn}"
    ws['A6'] = title
    ws['A6'].font = Font(name="Arial", bold=True, size=18, color="333333")
    ws['A6'].alignment = Alignment(horizontal='center')
    ws['A6'].border = Border(bottom=Side(style='medium', color=GOLD_HEX))

    ws.merge_cells('C8:D8')
    ws['C8'] = f"Date de facture: {data.get('date', '')}"
    ws['C8'].font = Font(name="Arial", size=10, color="555555")
    ws['C8'].alignment = Alignment(horizontal='right')

    headers = ['Description', 'Quantité', 'Prix DH', 'Montant DH']
    header_row = 10
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="333333")
        cell.fill = gray_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    items = data.get('items', [])
    for i, item in enumerate(items):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=item.get('description', '')).border = thin_border
        ws.cell(row=r, column=1).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=2, value=item.get('quantity', 0)).border = thin_border
        ws.cell(row=r, column=2).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=3, value=item.get('unit_price', 0)).border = thin_border
        ws.cell(row=r, column=3).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=4, value=item.get('total', 0)).border = thin_border
        ws.cell(row=r, column=4).font = Font(name="Arial", size=10)
        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')

    total = data.get('total', 0.0)
    remise = data.get('remise', 0.0)
    payer = data.get('payer', 0.0)
    net_total = data.get('net_total', total)

    r = header_row + 1 + len(items) + 1

    if remise > 0 or payer > 0:
        ws.cell(row=r, column=3, value="Total en DH").font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=total).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    if remise > 0:
        ws.cell(row=r, column=3, value="Remise en DH").font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=-remise).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    if payer > 0:
        ws.cell(row=r, column=3, value="Payer en DH").font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=-payer).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    ws.cell(row=r, column=3, value="Total à Payer en DH").font = Font(name="Arial", bold=True, size=11, color="333333")
    ws.cell(row=r, column=3).fill = gray_fill
    ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=3).border = thin_border
    ws.cell(row=r, column=4, value=net_total).font = Font(name="Arial", bold=True, size=12, color=GOLD_HEX)
    ws.cell(row=r, column=4).fill = gray_fill
    ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=4).border = thin_border

    base = safe_filename(data.get('client_name', ''))
    name = f"{base}.xlsx"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(tmp.name, as_attachment=True, download_name=name)


@invoices_bp.route('/preview', methods=['POST'])
@login_required
def preview():
    data = request.get_json(silent=True) or {}
    return render_template("invoice_pdf.html", data=data, company_name=COMPANY_NAME, company_address=COMPANY_ADDRESS)


@invoices_bp.route('/<int:invoice_id>/pdf')
@login_required
def download_saved_pdf(invoice_id):
    data = get_invoice_by_id(invoice_id)
    if not data:
        return "Facture non trouvée", 404

    pdf = InvoicePDF()
    pdf._client_name = data.get("client_name", "")
    pdf._date = data.get("date", "")
    pdf._facture_num = data.get("invoice_num", "")
    pdf._show_facture_num = bool(data.get("show_facture_num", False))
    pdf.set_auto_page_break(auto=True, margin=30)
    pdf.add_page()

    col_w = [72, 28, 44, 38]
    headers = ["Description", "Quantité", "Prix DH", "Montant DH"]

    def draw_table_header():
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_fill_color(*GRAY_BG)
        pdf.set_text_color(*DARK)
        pdf.set_draw_color(0, 0, 0)
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 8, h, border=1, align="C", fill=True)
        pdf.ln()

    draw_table_header()

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*DARK)
    last_page = pdf.page_no()
    for item in data.get('items', []):
        if pdf.page_no() != last_page:
            draw_table_header()
            last_page = pdf.page_no()
        pdf.cell(col_w[0], 7, item.get('description', ''), border=1)
        pdf.cell(col_w[1], 7, str(item.get('quantity', 0)), border=1, align="C")
        pdf.cell(col_w[2], 7, f"{item.get('unit_price', 0):,.2f}".replace(',', ' '), border=1, align="C")
        pdf.cell(col_w[3], 7, f"{item.get('total', 0):,.2f}".replace(',', ' '), border=1, align="C")
        pdf.ln()

    total = data.get('total', 0.0)
    remise = data.get('remise', 0.0)
    payer = data.get('payer', 0.0)
    net_total = data.get('net_total', total)

    pdf.ln(4)
    total_x = pdf.l_margin + col_w[0] + col_w[1]

    def price_line(label, value, val_color=None, bold=False):
        pdf.set_x(total_x)
        pdf.set_fill_color(*GRAY_BG)
        pdf.set_draw_color(0, 0, 0)
        pdf.set_font("Helvetica", "B" if bold else "", 10)
        pdf.set_text_color(*(val_color if val_color else DARK))
        pdf.cell(col_w[2], 10, label, border=1, align="C", fill=True)
        pdf.cell(col_w[3], 10, value, border=1, align="C", fill=True)
        pdf.ln(10)

    if remise > 0 or payer > 0:
        price_line("Total en DH", f"{total:,.2f} DH".replace(',', ' '))
    if remise > 0:
        price_line("Remise en DH", f"-{remise:,.2f} DH".replace(',', ' '))
    if payer > 0:
        price_line("Payer en DH", f"-{payer:,.2f} DH".replace(',', ' '))
    price_line("Total a Payer en DH", f"{net_total:,.2f} DH".replace(',', ' '), val_color=GOLD, bold=True)

    name = f"{safe_filename(data.get('client_name', ''))}.pdf"
    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=name, mimetype="application/pdf")


@invoices_bp.route('/<int:invoice_id>/excel')
@login_required
def download_saved_excel(invoice_id):
    data = get_invoice_by_id(invoice_id)
    if not data:
        return "Facture non trouvée", 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facture"

    GOLD_HEX = "C9A84C"
    GRAY_HEX = "F0F0F0"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color=GRAY_HEX, end_color=GRAY_HEX, fill_type="solid")

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    ws.merge_cells('A1:C1')
    ws['A1'] = "COCINA ESPAÑOLA"
    ws['A1'].font = Font(name="Arial", bold=True, size=20, color=GOLD_HEX)

    ws.merge_cells('A2:C2')
    ws['A2'] = "Art MDF"
    ws['A2'].font = Font(name="Arial", italic=True, size=14, color=GOLD_HEX)

    ws.merge_cells('D1:D2')
    ws['D1'] = "Rue Tange Center Al hirafiyin N 25"
    ws['D1'].font = Font(name="Arial", size=9, color="555555")
    ws['D1'].alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('D3:D3')
    ws['D3'] = "Imzouren AL Hoceima"
    ws['D3'].font = Font(name="Arial", size=9, color="555555")
    ws['D3'].alignment = Alignment(horizontal='right')

    ws['C4'] = "Client:"
    ws['C4'].font = Font(name="Arial", bold=True, size=10, color="333333")
    ws['C4'].alignment = Alignment(horizontal='right')

    ws['D4'] = data.get("client_name", "")
    ws['D4'].font = Font(name="Arial", bold=True, size=10, color="333333")
    ws['D4'].alignment = Alignment(horizontal='right')

    ws.merge_cells('A6:D6')
    title = "FACTURE"
    fn = data.get("invoice_num", "")
    if fn:
        title += f" N\u00b0 {fn}"
    ws['A6'] = title
    ws['A6'].font = Font(name="Arial", bold=True, size=18, color="333333")
    ws['A6'].alignment = Alignment(horizontal='center')
    ws['A6'].border = Border(bottom=Side(style='medium', color=GOLD_HEX))

    ws.merge_cells('C8:D8')
    ws['C8'] = f"Date de facture: {data.get('date', '')}"
    ws['C8'].font = Font(name="Arial", size=10, color="555555")
    ws['C8'].alignment = Alignment(horizontal='right')

    headers = ['Description', 'Quantité', 'Prix DH', 'Montant DH']
    header_row = 10
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="333333")
        cell.fill = gray_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    items = data.get('items', [])
    for i, item in enumerate(items):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=item.get('description', '')).border = thin_border
        ws.cell(row=r, column=1).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=2, value=item.get('quantity', 0)).border = thin_border
        ws.cell(row=r, column=2).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=3, value=item.get('unit_price', 0)).border = thin_border
        ws.cell(row=r, column=3).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=4, value=item.get('total', 0)).border = thin_border
        ws.cell(row=r, column=4).font = Font(name="Arial", size=10)
        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center')

    total = data.get('total', 0.0)
    remise = data.get('remise', 0.0)
    payer = data.get('payer', 0.0)
    net_total = data.get('net_total', total)

    r = header_row + 1 + len(items) + 1

    if remise > 0 or payer > 0:
        ws.cell(row=r, column=3, value="Total en DH").font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=total).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    if remise > 0:
        ws.cell(row=r, column=3, value="Remise en DH").font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=-remise).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    if payer > 0:
        ws.cell(row=r, column=3, value="Payer en DH").font = Font(name="Arial", bold=True, size=11, color="333333")
        ws.cell(row=r, column=3).fill = gray_fill
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=-payer).font = Font(name="Arial", size=12)
        ws.cell(row=r, column=4).fill = gray_fill
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=4).border = thin_border
        r += 1

    ws.cell(row=r, column=3, value="Total à Payer en DH").font = Font(name="Arial", bold=True, size=11, color="333333")
    ws.cell(row=r, column=3).fill = gray_fill
    ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=3).border = thin_border
    ws.cell(row=r, column=4, value=net_total).font = Font(name="Arial", bold=True, size=12, color=GOLD_HEX)
    ws.cell(row=r, column=4).fill = gray_fill
    ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=r, column=4).border = thin_border

    base = safe_filename(data.get('client_name', ''))
    name = f"{base}.xlsx"
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(tmp.name, as_attachment=True, download_name=name)


@invoices_bp.route('/<int:invoice_id>/delete', methods=['POST'])
@login_required
@idempotent_route()
def delete_invoice_route(invoice_id):
    try:
        invoice = db.session.get(Invoice, invoice_id)
        if invoice:
            db.session.delete(invoice)
            db.session.commit()
        return jsonify({"success": True})
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Erreur lors de la suppression."}), 500
